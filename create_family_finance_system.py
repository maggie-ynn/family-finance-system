import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建家庭财务管理系统
print("正在创建家庭财务管理系统...")

# 创建工作簿
wb = openpyxl.Workbook()

# 删除默认工作表
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# 定义样式
header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alignment_center = Alignment(horizontal='center', vertical='center')
alignment_left = Alignment(horizontal='left', vertical='center')

# 创建工作表列表
sheets_info = {
    "账户入金": {
        "headers": ["入金时间", "资金来源", "存入银行", "入金金额", "是否有支撑材料", "备注", "累计入金"],
        "widths": [18, 20, 18, 15, 18, 25, 15],
        "has_formula": True,
        "formula_col": 6  # 累计入金列索引
    },
    "贷款还款": {
        "headers": ["还款类型", "还款日期", "还款金额", "贷款类型", "期数", "本月利息", "本月本金", "备注", "累计还款"],
        "widths": [12, 15, 15, 15, 10, 15, 15, 25, 15],
        "has_formula": True,
        "formula_col": 8
    },
    "报税记录": {
        "headers": ["报税年份", "报税日期", "申报收入", "应纳税所得额", "报税金额", "实际缴税金额", "补缴/退税", "申报状态", "附件"],
        "widths": [12, 15, 15, 18, 15, 15, 15, 12, 18],
        "has_formula": True,
        "formula_col": 5  # 补缴/退税 = 实际缴税 - 报税金额
    },
    "免税账户管理": {
        "headers": ["账户名称", "银行名称", "账户类型", "账户余额", "年度投资收益", "年度取款", "剩余额度", "开户日期", "状态"],
        "widths": [20, 18, 15, 15, 18, 15, 15, 15, 12],
        "has_formula": True,
        "formula_col": 6  # 剩余额度
    },
    "教育账户管理": {
        "headers": ["学生姓名", "账户名称", "银行名称", "账户余额", "年度存入", "年度支取", "教育阶段", "开户日期", "备注"],
        "widths": [15, 20, 18, 15, 15, 15, 12, 15, 25],
        "has_formula": False
    },
    "收入支出跟踪": {
        "headers": ["交易日期", "交易类型", "收支类别", "金额", "账户", "交易对象", "项目/描述", "凭证附件", "是否分期", "分期数"],
        "widths": [15, 12, 15, 15, 18, 20, 25, 15, 12, 10],
        "has_formula": False
    }
}

# 创建每个工作表
for sheet_name, sheet_config in sheets_info.items():
    ws = wb.create_sheet(title=sheet_name)
    headers = sheet_config["headers"]
    widths = sheet_config["widths"]
    
    # 设置表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = alignment_center
    
    # 设置列宽
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 如果需要公式,在第二行设置示例公式
    if sheet_config["has_formula"]:
        formula_col = sheet_config["formula_col"]
        col_letter = get_column_letter(formula_col + 1)  # +1 因为 Excel 是从A开始
        
        if sheet_name == "账户入金":
            # 累计入金 = 当前行入金 + 上一行累计
            cell = ws.cell(row=2, column=formula_col + 1)
            cell.value = f"=IF(E2=\"\",0,E2) + IF(ROW()>2,{col_letter}1,0)"
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = alignment_center
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
        elif sheet_name == "贷款还款":
            # 累计还款 = 当行金额 + 上一行累计
            cell = ws.cell(row=2, column=formula_col + 1)
            cell.value = f"=IF(C2=\"\",0,C2) + IF(ROW()>2,{col_letter}1,0)"
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = alignment_center
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
        elif sheet_name == "报税记录":
            # 补缴/退税 = 实际缴税 - 报税金额
            cell = ws.cell(row=2, column=formula_col + 1)
            cell.value = "=IF(E2=\"\",0,E2) - IF(F2=\"\",0,F2)"
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = alignment_center
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
        elif sheet_name == "免税账户管理":
            # 剩余额度示例公式（需要根据实际政策调整）
            cell = ws.cell(row=2, column=formula_col + 1)
            cell.value = "=D2"  # 简化：剩余额度 = 账户余额（实际应根据年度投资额度计算）
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = alignment_center
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # 添加数据验证说明行（第3行）
    ws.row_dimensions[3].height = 30
    note_cell = ws.cell(row=3, column=1, value="说明：直接在此表格录入数据")
    note_cell.font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    note_cell.alignment = alignment_left
    
    # 冻结首行
    ws.freeze_panes = "A2"

# 创建仪表盘工作表
dashboard_ws = wb.create_sheet(title="财务仪表盘", index=0)
dashboard_ws.column_dimensions['A'].width = 25
dashboard_ws.column_dimensions['B'].width = 20
dashboard_ws.column_dimensions['C'].width = 25

# 仪表盘标题
title_cell = dashboard_ws.cell(row=1, column=1, value="家庭财务概览")
title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='4472C4')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
dashboard_ws.merge_cells('A1:C1')

# 仪表盘分类
categories = [
    ("资金流入", "账户入金"),
    ("负债管理", "贷款还款"),
    ("税务管理", "报税记录"),
    ("免税账户", "免税账户管理"),
    ("教育基金", "教育账户管理"),
    ("收支明细", "收入支出跟踪")
]

# 创建仪表盘链接
row = 3
for category, sheet_name in categories:
    # 类别名称
    cat_cell = dashboard_ws.cell(row=row, column=1, value=category)
    cat_cell.font = Font(name='微软雅黑', size=11, bold=True)
    cat_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cat_cell.border = border_thin
    cat_cell.alignment = alignment_center
    
    # 跳转链接
    link_cell = dashboard_ws.cell(row=row, column=2, value="点击查看明细")
    link_cell.font = Font(name='微软雅黑', size=10, color='4472C4', underline='single')
    link_cell.hyperlink = f"#'{sheet_name}'!A1"
    link_cell.border = border_thin
    link_cell.alignment = alignment_center
    
    # 记录数统计
    count_cell = dashboard_ws.cell(row=row, column=3, value=f"=COUNTA('{sheet_name}'!A:A)-1")
    count_cell.font = Font(name='微软雅黑', size=10)
    count_cell.border = border_thin
    count_cell.alignment = alignment_center
    
    row += 1

# 添加汇总统计区域
row += 2
summary_title = dashboard_ws.cell(row=row, column=1, value="关键指标汇总")
summary_title.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
summary_title.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
summary_title.alignment = alignment_center
dashboard_ws.merge_cells(f'A{row}:C{row}')
row += 1

# 汇总指标
metrics = [
    ("总入金金额", "=SUM('账户入金'!E:E)"),
    ("总还款金额", "=SUM('贷款还款'!C:C)"),
    ("总报税金额", "=SUM('报税记录'!E:E)"),
    ("总缴税金额", "=SUM('报税记录'!F:F)"),
    ("免税账户余额", "=SUM('免税账户管理'!D:D)"),
    ("教育账户余额", "=SUM('教育账户管理'!D:D)")
]

for metric_name, formula in metrics:
    name_cell = dashboard_ws.cell(row=row, column=1, value=metric_name)
    name_cell.font = Font(name='微软雅黑', size=10)
    name_cell.border = border_thin
    name_cell.alignment = alignment_left
    
    value_cell = dashboard_ws.cell(row=row, column=2, value=formula)
    value_cell.font = Font(name='微软雅黑', size=10, bold=True)
    value_cell.border = border_thin
    value_cell.alignment = alignment_center
    
    # 空列
    empty_cell = dashboard_ws.cell(row=row, column=3)
    empty_cell.border = border_thin
    
    row += 1

# 设置仪表盘行高
for i in range(1, row):
    dashboard_ws.row_dimensions[i].height = 25

# 保存文件
filename = "家庭财务管理系统.xlsx"
wb.save(filename)
wb.close()

print(f"✓ 家庭财务管理系统创建完成: {filename}")
print("\n系统包含以下模块:")
print("1. 账户入金 - 记录所有资金流入")
print("2. 贷款还款 - 跟踪各类贷款还款情况")
print("3. 报税记录 - 管理年度税务信息")
print("4. 免税账户管理 - TFSA等免税账户")
print("5. 教育账户管理 - RESP等教育基金")
print("6. 收入支出跟踪 - 日常收支明细")
print("7. 财务仪表盘 - 一览全局的关键指标")
print("\n使用说明:")
print("- 直接在各工作表的第3行开始录入数据")
print("- 第2行的公式会自动计算累计值")
print("- 仪表盘会自动汇总所有数据")
print("- 点击仪表盘中的链接可快速跳转到对应工作表")
