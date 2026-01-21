"""
家庭财务管理系统 - Excel 与网页同步工具

功能：
1. 从 Excel 读取数据并同步到网页
2. 从网页数据导出回 Excel
3. 支持增量同步和全量同步
4. 自动备份原始数据
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import os
from datetime import datetime
import shutil


class FinanceDataSync:
    def __init__(self, excel_path='家庭财务管理系统.xlsx', html_path='family_finance_web.html'):
        self.excel_path = excel_path
        self.html_path = html_path
        self.backup_path = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 工作表映射
        self.sheet_mapping = {
            'deposit': '账户入金',
            'loan': '贷款还款',
            'tax': '报税记录',
            'tfsa': '免税账户管理',
            'education': '教育账户管理',
            'expense': '收入支出跟踪'
        }
        
        # 字段映射（Excel 列名 -> 数据库字段名）
        self.field_mapping = {
            'deposit': {
                '入金时间': 'date',
                '资金来源': 'source',
                '存入银行': 'bank',
                '入金金额': 'amount',
                '是否有支撑材料': 'hasDocument',
                '备注': 'note'
            },
            'loan': {
                '还款类型': 'type',
                '还款日期': 'date',
                '还款金额': 'amount',
                '贷款类型': 'loanType',
                '期数': 'period',
                '本月利息': 'interest',
                '本月本金': 'principal',
                '备注': 'note'
            },
            'tax': {
                '报税年份': 'year',
                '报税日期': 'date',
                '申报收入': 'income',
                '应纳税所得额': 'taxableIncome',
                '报税金额': 'taxAmount',
                '实际缴税金额': 'paidAmount',
                '补缴/退税': 'diff',
                '申报状态': 'status',
                '附件': 'attachment'
            },
            'tfsa': {
                '账户名称': 'accountName',
                '银行名称': 'bank',
                '账户类型': 'accountType',
                '账户余额': 'balance',
                '年度投资收益': 'annualReturn',
                '年度取款': 'annualWithdrawal',
                '剩余额度': 'remaining',
                '开户日期': 'openDate',
                '状态': 'status'
            },
            'education': {
                '学生姓名': 'studentName',
                '账户名称': 'accountName',
                '银行名称': 'bank',
                '账户余额': 'balance',
                '年度存入': 'annualDeposit',
                '年度支取': 'annualWithdrawal',
                '教育阶段': 'educationStage',
                '开户日期': 'openDate',
                '备注': 'note'
            },
            'expense': {
                '交易日期': 'date',
                '交易类型': 'type',
                '收支类别': 'category',
                '金额': 'amount',
                '账户': 'account',
                '交易对象': 'counterparty',
                '项目/描述': 'description',
                '凭证附件': 'attachment',
                '是否分期': 'isInstallment',
                '分期数': 'installments'
            }
        }
    
    def backup_excel(self):
        """备份 Excel 文件"""
        if os.path.exists(self.excel_path):
            shutil.copy2(self.excel_path, self.backup_path)
            print(f"✓ 已备份原始文件: {self.backup_path}")
            return True
        return False
    
    def read_excel_data(self):
        """从 Excel 读取所有工作表数据"""
        if not os.path.exists(self.excel_path):
            print(f"✗ Excel 文件不存在: {self.excel_path}")
            return None
        
        print(f"正在读取 Excel 文件: {self.excel_path}")
        
        data = {
            'deposit': [],
            'loan': [],
            'tax': [],
            'tfsa': [],
            'education': [],
            'expense': []
        }
        
        try:
            wb = load_workbook(self.excel_path)
            
            for key, sheet_name in self.sheet_mapping.items():
                if sheet_name not in wb.sheetnames:
                    print(f"⚠ 工作表不存在: {sheet_name}")
                    continue
                
                ws = wb[sheet_name]
                field_map = self.field_mapping[key]
                
                # 获取表头
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(cell.value)
                
                # 读取数据（从第4行开始，跳过公式行和说明行）
                for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                    if not any(row):
                        continue
                    
                    record = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row) and header in field_map:
                            value = row[col_idx]
                            # 处理布尔值
                            if header in ['是否有支撑材料', '是否分期']:
                                value = bool(value) if value else False
                            # 处理金额
                            elif header in ['入金金额', '还款金额', '账户余额', '金额', 
                                          '申报收入', '报税金额', '实际缴税金额',
                                          '年度投资收益', '年度取款', '年度存入', '年度支取']:
                                try:
                                    value = float(value) if value else 0.0
                                except:
                                    value = 0.0
                            
                            record[field_map[header]] = value
                    
                    if record:
                        data[key].append(record)
                
                print(f"✓ 读取 {sheet_name}: {len(data[key])} 条记录")
            
            wb.close()
            return data
            
        except Exception as e:
            print(f"✗ 读取 Excel 失败: {str(e)}")
            return None
    
    def write_html_data(self, data):
        """将数据写入 HTML 文件"""
        if not os.path.exists(self.html_path):
            print(f"✗ HTML 文件不存在: {self.html_path}")
            return False
        
        print(f"正在同步数据到网页: {self.html_path}")
        
        try:
            # 读取原 HTML
            with open(self.html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 将数据转换为 JSON 字符串
            data_json = json.dumps(data, ensure_ascii=False, indent=2)
            
            # 查找并替换初始化数据
            old_pattern = r'let financeData = \{[^}]*\};'
            new_init = f'let financeData = {data_json};'
            
            import re
            new_html = re.sub(old_pattern, new_init, html_content, count=1, flags=re.DOTALL)
            
            # 写回文件
            with open(self.html_path, 'w', encoding='utf-8') as f:
                f.write(new_html)
            
            print(f"✓ 数据已同步到网页")
            print(f"  - 账户入金: {len(data.get('deposit', []))} 条")
            print(f"  - 贷款还款: {len(data.get('loan', []))} 条")
            print(f"  - 报税记录: {len(data.get('tax', []))} 条")
            print(f"  - 免税账户: {len(data.get('tfsa', []))} 条")
            print(f"  - 教育账户: {len(data.get('education', []))} 条")
            print(f"  - 收支跟踪: {len(data.get('expense', []))} 条")
            
            return True
            
        except Exception as e:
            print(f"✗ 同步到网页失败: {str(e)}")
            return False
    
    def export_to_excel(self, data):
        """将数据导出回 Excel"""
        print(f"正在导出数据到 Excel: {self.excel_path}")
        
        try:
            # 加载现有工作簿
            wb = load_workbook(self.excel_path)
            
            for key, sheet_name in self.sheet_mapping.items():
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                field_map = self.field_mapping[key]
                
                # 反向映射：字段名 -> Excel 列名
                reverse_map = {v: k for k, v in field_map.items()}
                
                # 获取表头
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(cell.value)
                
                # 获取列索引映射
                col_indices = {}
                for col_idx, header in enumerate(headers):
                    col_indices[header] = col_idx + 1  # Excel 列从 1 开始
                
                # 清空原有数据（保留前3行：表头、公式、说明）
                max_row = ws.max_row
                if max_row > 3:
                    for row_idx in range(max_row, 3, -1):
                        ws.delete_rows(row_idx)
                
                # 写入新数据（从第4行开始）
                records = data.get(key, [])
                for record_idx, record in enumerate(records, start=4):
                    for field_name, value in record.items():
                        if field_name in reverse_map:
                            header = reverse_map[field_name]
                            if header in col_indices:
                                cell = ws.cell(row=record_idx, column=col_indices[header])
                                cell.value = value
                                
                                # 设置格式
                                if header in ['是否有支撑材料', '是否分期']:
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                                elif header in ['入金金额', '还款金额', '账户余额', '金额',
                                              '申报收入', '报税金额', '实际缴税金额']:
                                    cell.number_format = '#,##0.00'
            
            # 保存
            wb.save(self.excel_path)
            wb.close()
            
            print(f"✓ 数据已导出到 Excel")
            return True
            
        except Exception as e:
            print(f"✗ 导出到 Excel 失败: {str(e)}")
            return False
    
    def excel_to_web(self):
        """Excel -> 网页同步"""
        print("\n" + "="*50)
        print("开始 Excel -> 网页 同步")
        print("="*50)
        
        self.backup_excel()
        data = self.read_excel_data()
        
        if data:
            self.write_html_data(data)
            print("\n✓ Excel -> 网页 同步完成")
        else:
            print("\n✗ Excel -> 网页 同步失败")
    
    def web_to_excel(self):
        """网页 -> Excel 同步"""
        print("\n" + "="*50)
        print("开始 网页 -> Excel 同步")
        print("="*50)
        
        self.backup_excel()
        
        # 从 HTML 读取数据
        try:
            with open(self.html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 提取 financeData
            import re
            pattern = r'let financeData = (\{.*?\});'
            match = re.search(pattern, html_content, re.DOTALL)
            
            if match:
                data_json = match.group(1)
                data = json.loads(data_json)
                
                if self.export_to_excel(data):
                    print("\n✓ 网页 -> Excel 同步完成")
                else:
                    print("\n✗ 网页 -> Excel 同步失败")
            else:
                print("✗ 无法从网页提取数据")
                
        except Exception as e:
            print(f"✗ 读取网页数据失败: {str(e)}")


def main():
    sync = FinanceDataSync()
    
    print("家庭财务管理系统 - 数据同步工具")
    print("="*50)
    print("1. Excel -> 网页（将 Excel 数据同步到网页）")
    print("2. 网页 -> Excel（将网页数据导出回 Excel）")
    print("3. 双向同步（先从网页读取，再从 Excel 读取）")
    print("="*50)
    
    choice = input("请选择同步方向（1/2/3，默认为1）: ").strip() or "1"
    
    if choice == "1":
        sync.excel_to_web()
    elif choice == "2":
        sync.web_to_excel()
    elif choice == "3":
        sync.web_to_excel()
        sync.excel_to_web()
    else:
        print("无效选择")


if __name__ == "__main__":
    main()
